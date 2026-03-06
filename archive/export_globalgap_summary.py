#!/usr/bin/env python3
"""Export GLOBALG.A.P. to readable formats"""

from openpyxl import load_workbook, Workbook
import csv

filename = '240902_IFA_GFS_checklist_FV_v6_0-GFS_Aug24_protected_en.xlsx'

try:
    wb = load_workbook(filename, read_only=True, data_only=True)
    ws = wb['PI']

    print("Reading GLOBALG.A.P. control points...")

    control_points = []

    for row_num in range(2, ws.max_row + 1):
        code = ws.cell(row=row_num, column=3).value
        if not code or not isinstance(code, str):
            continue

        control_points.append({
            'Code': code,
            'Section': ws.cell(row=row_num, column=15).value or '',
            'Subsection': ws.cell(row=row_num, column=19).value or '',
            'Control Point': ws.cell(row=row_num, column=5).value or '',
            'Compliance Criteria': ws.cell(row=row_num, column=7).value or '',
            'Criticality': ws.cell(row=row_num, column=9).value or ''
        })

    wb.close()

    # Export to CSV
    csv_file = 'GLOBALG.A.P._Control_Points.csv'
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        if control_points:
            writer = csv.DictWriter(f, fieldnames=control_points[0].keys())
            writer.writeheader()
            writer.writerows(control_points)
    print(f"✓ Exported to CSV: {csv_file}")

    # Export to clean Excel
    excel_file = 'GLOBALG.A.P._Control_Points_Clean.xlsx'
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Control Points"

    # Write headers
    headers = list(control_points[0].keys())
    for col_num, header in enumerate(headers, 1):
        ws_out.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, cp in enumerate(control_points, 2):
        for col_num, key in enumerate(headers, 1):
            ws_out.cell(row=row_num, column=col_num, value=cp[key])

    # Set column widths
    ws_out.column_dimensions['A'].width = 15  # Code
    ws_out.column_dimensions['B'].width = 35  # Section
    ws_out.column_dimensions['C'].width = 35  # Subsection
    ws_out.column_dimensions['D'].width = 70  # Control Point
    ws_out.column_dimensions['E'].width = 70  # Compliance Criteria
    ws_out.column_dimensions['F'].width = 15  # Criticality

    wb_out.save(excel_file)
    print(f"✓ Exported to Excel: {excel_file}")

    # Create text summary
    txt_file = 'GLOBALG.A.P._Summary.txt'
    with open(txt_file, 'w', encoding='utf-8') as f:
        f.write("="*100 + "\n")
        f.write("GLOBALG.A.P. IFA v6.0 - Fruit & Vegetables\n")
        f.write("Complete Control Points Summary\n")
        f.write("="*100 + "\n\n")

        # Group by section
        sections = {}
        for cp in control_points:
            section = cp['Section']
            if section not in sections:
                sections[section] = []
            sections[section].append(cp)

        for section_name in sorted(sections.keys()):
            f.write(f"\n{'='*100}\n")
            f.write(f"{section_name}\n")
            f.write(f"{'='*100}\n\n")

            for i, cp in enumerate(sections[section_name], 1):
                f.write(f"{i}. {cp['Code']} - {cp['Criticality']}\n")
                f.write(f"   {cp['Control Point']}\n\n")

        f.write(f"\n{'='*100}\n")
        f.write(f"TOTAL: {len(control_points)} Control Points\n")
        f.write(f"{'='*100}\n")

    print(f"✓ Exported to TXT: {txt_file}")

    # Print summary
    print(f"\n{'='*80}")
    print(f"EXPORT COMPLETE")
    print(f"{'='*80}\n")
    print(f"Total Control Points: {len(control_points)}")

    sections_count = {}
    for cp in control_points:
        section = cp['Section']
        sections_count[section] = sections_count.get(section, 0) + 1

    print(f"\nSections ({len(sections_count)}):")
    for section, count in sorted(sections_count.items()):
        print(f"  • {section}: {count} points")

    major = len([cp for cp in control_points if 'Major' in cp['Criticality']])
    minor = len([cp for cp in control_points if 'Minor' in cp['Criticality']])

    print(f"\nCriticality:")
    print(f"  🔴 Major Must: {major}")
    print(f"  🟠 Minor Must: {minor}")

    print(f"\n{'='*80}")
    print(f"Files created:")
    print(f"  1. {csv_file} (for Excel/import)")
    print(f"  2. {excel_file} (clean format)")
    print(f"  3. {txt_file} (readable summary)")
    print(f"{'='*80}\n")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
