#!/usr/bin/env python3
"""Inspect the control point sheets in the GLOBALG.A.P. checklist"""

from openpyxl import load_workbook

filename = '240902_IFA_GFS_checklist_FV_v6_0-GFS_Aug24_protected_en.xlsx'

try:
    wb = load_workbook(filename, read_only=True, data_only=True)

    # Check the main sheets that likely contain control points
    sheets_to_check = ['PI', 'S', 'PQ', 'Sheet1']

    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        print(f"\n{'='*80}")
        print(f"Sheet: '{sheet_name}' ({ws.max_row} rows, {ws.max_column} columns)")
        print(f"{'='*80}")

        # Find header row (usually has "Control Point" or "CP" text)
        header_row = None
        for row_num in range(1, min(20, ws.max_row + 1)):
            for col_num in range(1, min(20, ws.max_column + 1)):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value and isinstance(cell_value, str):
                    if 'control point' in cell_value.lower() or 'compliance criteria' in cell_value.lower():
                        header_row = row_num
                        break
            if header_row:
                break

        if not header_row:
            header_row = 1

        print(f"\nAssumed header row: {header_row}")
        print(f"\nHeaders:")
        headers = {}
        for col_num in range(1, min(30, ws.max_column + 1)):
            cell = ws.cell(row=header_row, column=col_num)
            if cell.value:
                headers[col_num] = str(cell.value)
                print(f"  Col {col_num:2d}: {cell.value}")

        # Show a few sample data rows
        print(f"\nSample data rows (showing first 3 data rows):")
        data_start_row = header_row + 1
        for row_num in range(data_start_row, min(data_start_row + 3, ws.max_row + 1)):
            print(f"\n  Row {row_num}:")
            for col_num in headers.keys():
                cell = ws.cell(row=row_num, column=col_num)
                value = str(cell.value)[:60] if cell.value else ""
                if value and value != "None":
                    print(f"    {headers[col_num]}: {value}")

    wb.close()

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
