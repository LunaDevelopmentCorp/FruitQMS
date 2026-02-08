#!/usr/bin/env python3
"""
Import Official GLOBALG.A.P. IFA v6.0 Checklist
Specifically designed for the file: 240902_IFA_GFS_checklist_FV_v6_0-GFS_Aug24_protected_en.xlsx
"""

from openpyxl import load_workbook
from app import create_app, db
from app.models import Organization, ControlPoint
import json

def import_official_globalgap(filename, organization_id=None):
    """Import control points from official GLOBALG.A.P. Excel file"""

    app = create_app()

    with app.app_context():
        try:
            print(f"\n{'='*80}")
            print(f"Importing Official GLOBALG.A.P. IFA v6.0 Checklist")
            print(f"{'='*80}\n")

            wb = load_workbook(filename, read_only=True, data_only=True)

            # Get or prompt for organization
            if organization_id:
                org = Organization.query.get(organization_id)
            else:
                orgs = Organization.query.all()
                if len(orgs) == 0:
                    print("❌ Error: No organizations found. Complete the setup wizard first.")
                    return
                elif len(orgs) == 1:
                    org = orgs[0]
                    print(f"✓ Using organization: {org.name}\n")
                else:
                    print("\nAvailable organizations:")
                    for i, o in enumerate(orgs, 1):
                        print(f"  {i}. {o.name} (ID: {o.id}, Type: {o.org_type})")
                    choice = int(input("\n👉 Select organization number: ")) - 1
                    org = orgs[choice]
                    print(f"\n✓ Selected: {org.name}\n")

            if not org:
                print("❌ Error: Organization not found")
                return

            # Read from PI sheet (Product Integrity control points)
            ws = wb['PI']

            control_points = []
            existing_codes = set()

            print("📋 Reading control points from 'PI' sheet...")

            # Column mapping (based on inspection):
            # Col 3: Number (e.g., FV-GFS 10.01)
            # Col 5: P (Control Point text)
            # Col 7: C (Compliance Criteria text)
            # Col 9: L (Criticality: Major Must, Minor Must, Recommendation)
            # Col 15: S (Section name, e.g., "FV 10 COMPLAINTS")
            # Col 19: SS (Subsection name)
            # Col 17: Order (Section number)

            for row_num in range(2, ws.max_row + 1):
                code = ws.cell(row=row_num, column=3).value  # Number
                if not code or not isinstance(code, str):
                    continue

                # Skip duplicates
                if code in existing_codes:
                    continue
                existing_codes.add(code)

                control_point_text = ws.cell(row=row_num, column=5).value or ''  # P
                compliance_criteria = ws.cell(row=row_num, column=7).value or ''  # C
                criticality = ws.cell(row=row_num, column=9).value or 'Major Must'  # L
                section = ws.cell(row=row_num, column=15).value or ''  # S
                subsection = ws.cell(row=row_num, column=19).value or ''  # SS
                order = ws.cell(row=row_num, column=17).value or 0  # Order

                # Determine category from code
                if code.startswith('FV-GFS'):
                    category = 'Fruit & Vegetables'
                elif code.startswith('AF-GFS'):
                    category = 'All Farm Base'
                else:
                    category = 'General'

                # Determine applies_to based on section
                section_lower = section.lower() if section else ''
                if 'produce handling' in section_lower or 'packhouse' in section_lower:
                    applies_to = 'Packhouse'
                elif 'harvest' in section_lower or 'crop' in section_lower or 'field' in section_lower:
                    applies_to = 'Grower'
                else:
                    applies_to = 'All'

                control_point = {
                    'code': code,
                    'category': category,
                    'section': section,
                    'subsection': subsection if subsection and subsection != '-' else '',
                    'control_point': control_point_text,
                    'compliance_criteria': compliance_criteria,
                    'criticality': criticality,
                    'level': 'L1',  # Default level
                    'applies_to': applies_to,
                    'order': order,
                    'overlap_schemes': 'GLOBALG.A.P. IFA v6.0'
                }

                control_points.append(control_point)

            wb.close()

            print(f"✓ Found {len(control_points)} unique control points\n")

            # Delete existing control points for this organization
            existing_count = ControlPoint.query.filter_by(organization_id=org.id).count()
            if existing_count > 0:
                print(f"⚠️  Found {existing_count} existing control points")
                response = input("   Delete and replace them? (yes/no): ").lower()
                if response == 'yes':
                    ControlPoint.query.filter_by(organization_id=org.id).delete()
                    db.session.commit()
                    print(f"✓ Deleted {existing_count} existing control points\n")
                else:
                    print("❌ Import cancelled")
                    return

            # Insert control points into database
            print("📥 Importing control points to database...")
            inserted_count = 0

            for cp_data in control_points:
                # Store FULL compliance criteria (no truncation)
                full_notes = f"Compliance Criteria: {cp_data['compliance_criteria']}"

                cp = ControlPoint(
                    organization_id=org.id,
                    code=cp_data['code'],
                    category=cp_data['category'],
                    description=f"{cp_data['section']} - {cp_data['control_point']}",
                    criticality=cp_data['criticality'],
                    compliance_status='N/A',  # Default status
                    overlap_hint=cp_data['overlap_schemes'],
                    notes=full_notes  # FULL compliance criteria stored here
                )
                db.session.add(cp)
                inserted_count += 1

                # Commit in batches
                if inserted_count % 50 == 0:
                    db.session.commit()
                    print(f"   Imported {inserted_count}/{len(control_points)} control points...")

            db.session.commit()

            print(f"\n{'='*80}")
            print(f"✅ IMPORT COMPLETE!")
            print(f"{'='*80}\n")

            print(f"Summary:")
            print(f"  • Organization: {org.name}")
            print(f"  • Total control points: {len(control_points)}")

            # Count by criticality
            major = len([cp for cp in control_points if 'Major' in cp['criticality']])
            minor = len([cp for cp in control_points if 'Minor' in cp['criticality']])
            rec = len([cp for cp in control_points if 'Recommendation' in cp['criticality']])

            print(f"\n  By Criticality:")
            print(f"    🔴 Major Must: {major}")
            print(f"    🟠 Minor Must: {minor}")
            print(f"    🟢 Recommendations: {rec}")

            # Count by category
            categories = {}
            for cp in control_points:
                cat = cp['category']
                categories[cat] = categories.get(cat, 0) + 1

            print(f"\n  By Category:")
            for category, count in sorted(categories.items()):
                print(f"    • {category}: {count} points")

            # Count by applies_to
            applies = {}
            for cp in control_points:
                app_to = cp['applies_to']
                applies[app_to] = applies.get(app_to, 0) + 1

            print(f"\n  By Applicability:")
            for app_to, count in sorted(applies.items()):
                print(f"    • {app_to}: {count} points")

            print(f"\n{'='*80}")
            print(f"✅ Control points are now in FruitQMS!")
            print(f"{'='*80}\n")
            print(f"Next steps:")
            print(f"  1. Go to: http://127.0.0.1:5000/dashboard")
            print(f"  2. Click: Settings > GLOBALG.A.P. Checklist")
            print(f"  3. Mark compliance status for each control point")
            print(f"  4. Upload evidence documents")
            print(f"  5. Track your certification progress!\n")

        except Exception as e:
            print(f"\n❌ Error importing: {str(e)}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        org_id = int(sys.argv[1]) if sys.argv[1].isdigit() else None
        import_official_globalgap('240902_IFA_GFS_checklist_FV_v6_0-GFS_Aug24_protected_en.xlsx', org_id)
    else:
        import_official_globalgap('240902_IFA_GFS_checklist_FV_v6_0-GFS_Aug24_protected_en.xlsx')
