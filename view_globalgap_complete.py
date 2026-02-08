#!/usr/bin/env python3
"""View complete contents of the GLOBALG.A.P. Excel file"""

from openpyxl import load_workbook
import json

filename = '240902_IFA_GFS_checklist_FV_v6_0-GFS_Aug24_protected_en.xlsx'

try:
    wb = load_workbook(filename, read_only=True, data_only=True)
    ws = wb['PI']  # Main control points sheet

    print(f"{'='*100}")
    print(f"GLOBALG.A.P. IFA v6.0 - Complete Control Points List")
    print(f"{'='*100}\n")

    control_points = []

    for row_num in range(2, ws.max_row + 1):
        code = ws.cell(row=row_num, column=3).value  # Number
        if not code or not isinstance(code, str):
            continue

        control_point_text = ws.cell(row=row_num, column=5).value or ''  # P
        compliance_criteria = ws.cell(row=row_num, column=7).value or ''  # C
        criticality = ws.cell(row=row_num, column=9).value or ''  # L
        section = ws.cell(row=row_num, column=15).value or ''  # S
        subsection = ws.cell(row=row_num, column=19).value or ''  # SS

        control_points.append({
            'code': code,
            'section': section,
            'subsection': subsection,
            'control_point': control_point_text,
            'compliance_criteria': compliance_criteria,
            'criticality': criticality
        })

    wb.close()

    # Group by section
    sections = {}
    for cp in control_points:
        section = cp['section']
        if section not in sections:
            sections[section] = []
        sections[section].append(cp)

    # Display organized by section
    for section_name in sorted(sections.keys()):
        print(f"\n{'='*100}")
        print(f"📋 {section_name}")
        print(f"{'='*100}\n")

        section_cps = sections[section_name]
        for i, cp in enumerate(section_cps, 1):
            print(f"{i}. [{cp['code']}] {cp['criticality']}")
            print(f"   Control Point: {cp['control_point'][:150]}{'...' if len(cp['control_point']) > 150 else ''}")
            print(f"   Compliance: {cp['compliance_criteria'][:150]}{'...' if len(cp['compliance_criteria']) > 150 else ''}")
            if cp['subsection'] and cp['subsection'] != '-':
                print(f"   Subsection: {cp['subsection']}")
            print()

    # Summary
    print(f"\n{'='*100}")
    print(f"SUMMARY")
    print(f"{'='*100}\n")
    print(f"Total Control Points: {len(control_points)}")
    print(f"Total Sections: {len(sections)}")

    # Count by criticality
    major = len([cp for cp in control_points if 'Major' in cp['criticality']])
    minor = len([cp for cp in control_points if 'Minor' in cp['criticality']])
    rec = len([cp for cp in control_points if 'Recommendation' in cp['criticality']])

    print(f"\nBy Criticality:")
    print(f"  🔴 Major Must: {major}")
    print(f"  🟠 Minor Must: {minor}")
    print(f"  🟢 Recommendations: {rec}")

    print(f"\nSections:")
    for section_name in sorted(sections.keys()):
        print(f"  • {section_name}: {len(sections[section_name])} points")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
