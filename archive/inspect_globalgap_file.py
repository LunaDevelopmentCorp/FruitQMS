#!/usr/bin/env python3
"""Inspect the structure of the official GLOBALG.A.P. checklist"""

from openpyxl import load_workbook

filename = '240902_IFA_GFS_checklist_FV_v6_0-GFS_Aug24_protected_en.xlsx'

try:
    wb = load_workbook(filename, read_only=True, data_only=True)

    print(f"✓ Successfully opened: {filename}\n")
    print(f"Sheet Names ({len(wb.sheetnames)} total):")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet_name}")

    # Inspect first sheet with data
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip empty sheets
        if ws.max_row < 2:
            continue

        print(f"\n{'='*80}")
        print(f"Inspecting sheet: '{sheet_name}'")
        print(f"{'='*80}")
        print(f"Max Row: {ws.max_row}")
        print(f"Max Column: {ws.max_column}")

        # Show first few rows
        print(f"\nFirst 5 rows:")
        for row_num in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                value = cell.value
                if value:
                    row_data.append(f"Col{col_num}={str(value)[:30]}")
            if row_data:
                print(f"  Row {row_num}: {', '.join(row_data)}")

        # Show headers (usually row 1 or 2)
        print(f"\nColumn Headers (Row 1):")
        for col_num in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=1, column=col_num)
            if cell.value:
                print(f"  Column {col_num}: {cell.value}")

        print(f"\nColumn Headers (Row 2, if different):")
        for col_num in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=2, column=col_num)
            if cell.value:
                print(f"  Column {col_num}: {cell.value}")

        # Only inspect first 2 sheets with data
        break

    wb.close()

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
