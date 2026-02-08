#!/usr/bin/env python3
"""
GLOBALG.A.P. IFA v6.0 Control Points Import Utility

This script imports control points from an Excel file into the FruitQMS database.
It can also generate a sample template Excel file.

Usage:
    python3 import_globalgap.py --template          # Generate Excel template
    python3 import_globalgap.py --import <file>     # Import from Excel file
    python3 import_globalgap.py --sample            # Generate sample with test data
"""

import sys
import argparse
from app import create_app, db
from app.models import Organization
import json

# Check if openpyxl is available
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)


def generate_template(filename='GLOBALG.A.P._Template.xlsx', include_sample=False):
    """Generate an Excel template for GLOBALG.A.P. control points"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Control Points"

    # Define headers
    headers = [
        'Code',
        'Section',
        'Category',
        'Sub-Category',
        'Control Point',
        'Compliance Criteria',
        'Criticality',
        'Level',
        'Applies To',
        'Overlap Schemes',
        'Notes'
    ]

    # Style headers
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 12  # Code
    ws.column_dimensions['B'].width = 20  # Section
    ws.column_dimensions['C'].width = 25  # Category
    ws.column_dimensions['D'].width = 25  # Sub-Category
    ws.column_dimensions['E'].width = 50  # Control Point
    ws.column_dimensions['F'].width = 60  # Compliance Criteria
    ws.column_dimensions['G'].width = 15  # Criticality
    ws.column_dimensions['H'].width = 10  # Level
    ws.column_dimensions['I'].width = 20  # Applies To
    ws.column_dimensions['J'].width = 25  # Overlap Schemes
    ws.column_dimensions['K'].width = 30  # Notes

    if include_sample:
        # Add sample data for reference
        sample_data = [
            ['AF 1.1.1', 'All Farm Base', 'Site Management', 'Site History',
             'Is there documented evidence that a risk assessment of the site history has been carried out?',
             'Records of risk assessment considering previous land use, potential contamination sources',
             'Major Must', 'L1', 'All', 'GRASP, SMETA', 'Essential for site approval'],

            ['AF 1.2.1', 'All Farm Base', 'Site Management', 'Site Management Plan',
             'Is there a documented site management plan?',
             'Written plan covering conservation areas, buffer zones, biodiversity',
             'Major Must', 'L1', 'All', 'Rainforest Alliance', 'Update annually'],

            ['AF 2.1.1', 'All Farm Base', 'Record Keeping & Self Assessment', 'Internal Inspections',
             'Has an internal inspection been carried out at least once per year?',
             'Records of internal inspections with dates, findings, and corrective actions',
             'Major Must', 'L1', 'All', 'GLOBALG.A.P.', 'Pre-audit requirement'],

            ['FV 1.1.1', 'Fruit & Vegetables', 'Propagation Material', 'Seed/Plant Source',
             'Are seeds/seedlings purchased from registered suppliers?',
             'Supplier invoices, plant passports, phytosanitary certificates',
             'Major Must', 'L1', 'Grower, Mixed', 'LEAF', 'Traceability requirement'],

            ['FV 1.2.1', 'Fruit & Vegetables', 'Crop Protection', 'IPM Program',
             'Is there a documented Integrated Pest Management (IPM) program?',
             'Written IPM strategy including pest monitoring, thresholds, biological control',
             'Major Must', 'L1', 'Grower, Mixed', 'LEAF, Tesco Nurture', 'Core requirement'],

            ['FV 2.1.1', 'Fruit & Vegetables', 'Produce Handling', 'Hygiene',
             'Is there a hygiene policy for the packhouse?',
             'Written hygiene procedures covering cleaning schedules, personnel hygiene, pest control',
             'Major Must', 'L1', 'Packhouse, Mixed', 'BRC, IFS', 'Food safety critical'],

            ['FV 2.2.1', 'Fruit & Vegetables', 'Produce Handling', 'Traceability',
             'Can produce be traced from packhouse back to field?',
             'Lot numbers, field codes, harvest dates linking product to origin',
             'Major Must', 'L1', 'Packhouse, Mixed', 'BRC, IFS, Tesco', 'One-up, one-down'],

            ['AF 3.1.1', 'All Farm Base', 'Workers Health Safety & Welfare', 'Risk Assessment',
             'Has a health and safety risk assessment been carried out?',
             'Documented risk assessment covering all work activities and hazards',
             'Major Must', 'L1', 'All', 'GRASP, SMETA', 'Legal requirement'],

            ['AF 4.1.1', 'All Farm Base', 'Waste & Pollution', 'Waste Management',
             'Is there a documented waste management plan?',
             'Written plan for waste segregation, recycling, disposal of all waste types',
             'Minor Must', 'L1', 'All', 'Rainforest Alliance', 'Environmental protection'],

            ['AF 5.1.1', 'All Farm Base', 'Environment & Biodiversity', 'Conservation',
             'Are conservation areas identified and protected?',
             'Maps showing conservation areas, buffer zones, and management practices',
             'Minor Must', 'L2', 'All', 'Rainforest Alliance', 'Biodiversity priority']
        ]

        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

    # Add instructions sheet
    ws_info = wb.create_sheet("Instructions")
    ws_info.column_dimensions['A'].width = 80

    instructions = [
        "GLOBALG.A.P. IFA v6.0 Control Points Import Template",
        "",
        "INSTRUCTIONS:",
        "1. Fill in the 'Control Points' sheet with your complete GLOBALG.A.P. scope",
        "2. Each row represents one control point",
        "",
        "COLUMN DESCRIPTIONS:",
        "• Code: Control point code (e.g., AF 1.1.1, FV 2.3.1)",
        "• Section: Main section (e.g., 'All Farm Base', 'Fruit & Vegetables')",
        "• Category: Category name (e.g., 'Site Management', 'Crop Protection')",
        "• Sub-Category: Sub-category if applicable",
        "• Control Point: The actual control point question/requirement",
        "• Compliance Criteria: What evidence is needed for compliance",
        "• Criticality: Must be one of: 'Major Must', 'Minor Must', 'Recommendation'",
        "• Level: Audit level (e.g., L1, L2, L1+L2)",
        "• Applies To: Who this applies to: 'All', 'Grower', 'Packhouse', 'Mixed'",
        "• Overlap Schemes: Other schemes this overlaps with (comma-separated)",
        "• Notes: Any additional notes or guidance",
        "",
        "CRITICALITY LEVELS:",
        "• Major Must (Red): Critical for certification, zero tolerance",
        "• Minor Must (Orange): Required for certification, some tolerance",
        "• Recommendation (Green): Best practice, not mandatory",
        "",
        "APPLIES TO OPTIONS:",
        "• All: Applies to all operations (growers and packhouses)",
        "• Grower: Only applies to growers/farms with own fields",
        "• Packhouse: Only applies to packhouse operations",
        "• Mixed: Applies to operations with both packhouse and own fields",
        "",
        "OVERLAP SCHEMES (examples):",
        "• GRASP (social compliance)",
        "• SMETA (ethical trade)",
        "• Rainforest Alliance",
        "• LEAF (integrated farm management)",
        "• Tesco Nurture",
        "• BRC, IFS (food safety)",
        "",
        "TO IMPORT:",
        "Run: python3 import_globalgap.py --import GLOBALG.A.P._Template.xlsx",
        "",
        "NOTE: You can export your current GLOBALG.A.P. Excel file directly,",
        "or use this template as a guide for formatting your data."
    ]

    for row_num, instruction in enumerate(instructions, 1):
        ws_info.cell(row=row_num, column=1, value=instruction)
        if row_num in [1, 7, 14, 19, 23]:  # Bold certain rows
            ws_info.cell(row=row_num, column=1).font = Font(bold=True, size=12)

    wb.save(filename)
    print(f"✓ Template saved to: {filename}")
    print(f"  - Open in Excel, fill in your control points")
    print(f"  - Then run: python3 import_globalgap.py --import {filename}")


def import_from_excel(filename, organization_id=None):
    """Import control points from Excel file into database"""

    app = create_app()

    with app.app_context():
        try:
            wb = load_workbook(filename)
            ws = wb['Control Points']

            # Get or prompt for organization
            if organization_id:
                org = Organization.query.get(organization_id)
            else:
                orgs = Organization.query.all()
                if len(orgs) == 0:
                    print("Error: No organizations found. Complete the setup wizard first.")
                    return
                elif len(orgs) == 1:
                    org = orgs[0]
                    print(f"Using organization: {org.name}")
                else:
                    print("\nAvailable organizations:")
                    for i, o in enumerate(orgs, 1):
                        print(f"  {i}. {o.name} (ID: {o.id})")
                    choice = int(input("\nSelect organization number: ")) - 1
                    org = orgs[choice]

            if not org:
                print("Error: Organization not found")
                return

            # Parse control points
            control_points = []
            row_num = 2  # Start after header

            while True:
                code = ws.cell(row=row_num, column=1).value
                if not code:
                    break

                control_point = {
                    'code': code,
                    'section': ws.cell(row=row_num, column=2).value or '',
                    'category': ws.cell(row=row_num, column=3).value or '',
                    'sub_category': ws.cell(row=row_num, column=4).value or '',
                    'control_point': ws.cell(row=row_num, column=5).value or '',
                    'compliance_criteria': ws.cell(row=row_num, column=6).value or '',
                    'criticality': ws.cell(row=row_num, column=7).value or 'Major Must',
                    'level': ws.cell(row=row_num, column=8).value or 'L1',
                    'applies_to': ws.cell(row=row_num, column=9).value or 'All',
                    'overlap_schemes': ws.cell(row=row_num, column=10).value or '',
                    'notes': ws.cell(row=row_num, column=11).value or ''
                }

                control_points.append(control_point)
                row_num += 1

            print(f"\n✓ Found {len(control_points)} control points in Excel file")

            # Store as JSON in organization
            org.intake_protocols = json.dumps({
                'globalgap_ifa_v6': {
                    'imported_at': str(db.func.now()),
                    'total_points': len(control_points),
                    'control_points': control_points
                }
            })

            db.session.commit()

            print(f"✓ Imported successfully into organization: {org.name}")
            print(f"\nSummary:")
            print(f"  - Total control points: {len(control_points)}")

            # Count by criticality
            major = len([cp for cp in control_points if 'Major' in cp['criticality']])
            minor = len([cp for cp in control_points if 'Minor' in cp['criticality']])
            rec = len([cp for cp in control_points if 'Recommendation' in cp['criticality']])

            print(f"  - Major Must: {major}")
            print(f"  - Minor Must: {minor}")
            print(f"  - Recommendations: {rec}")

            # Count by section
            sections = {}
            for cp in control_points:
                section = cp['section']
                sections[section] = sections.get(section, 0) + 1

            print(f"\n  By Section:")
            for section, count in sections.items():
                print(f"    • {section}: {count} points")

            print(f"\n✓ Import complete! Control points are now available in FruitQMS.")
            print(f"  Go to Settings > GLOBALG.A.P. Checklist to view and track compliance.")

        except Exception as e:
            print(f"Error importing: {str(e)}")
            import traceback
            traceback.print_exc()


def main():
    parser = argparse.ArgumentParser(
        description='GLOBALG.A.P. Control Points Import Utility',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 import_globalgap.py --template
      Generate a blank Excel template

  python3 import_globalgap.py --sample
      Generate an Excel template with sample data

  python3 import_globalgap.py --import myfile.xlsx
      Import control points from Excel file

  python3 import_globalgap.py --import myfile.xlsx --org 1
      Import for specific organization ID
        """
    )

    parser.add_argument('--template', action='store_true',
                       help='Generate blank Excel template')
    parser.add_argument('--sample', action='store_true',
                       help='Generate Excel template with sample data')
    parser.add_argument('--import', dest='import_file', metavar='FILE',
                       help='Import control points from Excel file')
    parser.add_argument('--org', type=int, metavar='ID',
                       help='Organization ID to import for')

    args = parser.parse_args()

    if args.template:
        generate_template(include_sample=False)
    elif args.sample:
        generate_template('GLOBALG.A.P._Sample.xlsx', include_sample=True)
    elif args.import_file:
        import_from_excel(args.import_file, args.org)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
